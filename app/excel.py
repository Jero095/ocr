"""Build .xlsx workbooks from parsed statements.

Amounts are written as real numbers with a number format, not as text, so the
figures are summable in Excel. Percentages are stored as fractions (5.00% ->
0.05) with a percent format, which is how Excel represents them natively.

An unambiguous mm/dd/yyyy date (4-digit year) is written as a real Excel date so
it sorts and filters properly. Anything less certain stays text: policy numbers
like 'AVWV009000353300', and two-digit-year dates like Chris Leef's '05/26/26',
where guessing the century or the day/month order would silently corrupt data.
"""

from __future__ import annotations

import re
from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .extract import CANONICAL_FIELDS, canonical_map, is_percent, to_float

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
TOTAL_FONT = Font(bold=True)
TOP_BORDER = Border(top=Side(style="thin", color="9CA3AF"))
OK_FONT = Font(bold=True, color="1A7F37")
BAD_FONT = Font(bold=True, color="CF222E")

MONEY_FMT = "#,##0.00"
PERCENT_FMT = "0.00%"
PLAIN_FMT = "0.00"
DATE_FMT = "mm/dd/yyyy"

# Only a 4-digit-year mm/dd/yyyy is written as a real Excel date. A 2-digit
# year (Chris Leef prints 05/26/26) stays text: guessing the century, or the
# day/month order, would silently alter data.
STRICT_MDY_RE = re.compile(r"^(\d{2})/(\d{2})/(\d{4})$")
PERCENT_HEADER_RE = re.compile(r"%|percent|\brate\b", re.I)

def _sheet_name(raw: str, used: set[str]) -> str:
    """Excel sheet names: <=31 chars, no []:*?/\\ , and unique in the book."""
    name = re.sub(r"[\[\]:*?/\\]", "-", raw).strip() or "Statement"
    name = name[:31]
    if name in used:
        stem = name[:28]
        n = 2
        while f"{stem}~{n}" in used:
            n += 1
        name = f"{stem}~{n}"
    used.add(name)
    return name


def _write(cell, text: str, header: str = "") -> None:
    """Write a cell as a typed number/date where unambiguous, else as text."""
    s = (text or "").strip()

    m = STRICT_MDY_RE.match(s)
    if m:
        month, day, year = (int(g) for g in m.groups())
        try:
            cell.value = date(year, month, day)
        except ValueError:
            cell.value = s
        else:
            cell.number_format = DATE_FMT
            cell.alignment = Alignment(horizontal="right")
        return

    value = to_float(s)
    if value is None:
        cell.value = text
        return
    if is_percent(s):
        cell.value = value / 100
        cell.number_format = PERCENT_FMT
    elif PERCENT_HEADER_RE.search(header):
        # A percent column whose '%' was stripped by the source export: keep the
        # number as-is rather than implying currency.
        cell.value = value
        cell.number_format = PLAIN_FMT
    else:
        cell.value = value
        cell.number_format = MONEY_FMT
    cell.alignment = Alignment(horizontal="right")


def _style_header(ws, ncols: int, row: int = 1) -> None:
    for i in range(1, ncols + 1):
        cell = ws.cell(row=row, column=i)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28


def _autosize(ws, headers: list[str], rows: list[list[str]]) -> None:
    for i, head in enumerate(headers, start=1):
        widest = len(str(head))
        for row in rows:
            if i - 1 < len(row):
                widest = max(widest, len(str(row[i - 1] or "")))
        ws.column_dimensions[get_column_letter(i)].width = min(max(widest + 3, 11), 42)


def _statement_sheet(wb: Workbook, stmt: dict, used: set[str]) -> None:
    ws = wb.create_sheet(_sheet_name(stmt["filename"].rsplit(".", 1)[0], used))
    cols = stmt["columns"]

    ws.append(cols)
    _style_header(ws, len(cols))

    for row in stmt["rows"]:
        ws.append([""] * len(cols))
        for i, col in enumerate(cols, start=1):
            _write(ws.cell(row=ws.max_row, column=i), row.get(col, ""), col)

    for sub in stmt.get("subtotals", []):
        ws.append([""] * len(cols))
        for i, col in enumerate(cols, start=1):
            cell = ws.cell(row=ws.max_row, column=i)
            _write(cell, sub.get(col, ""), col)
            cell.font = Font(italic=True)

    if stmt.get("totals"):
        ws.append([""] * len(cols))
        for i, col in enumerate(cols, start=1):
            cell = ws.cell(row=ws.max_row, column=i)
            _write(cell, stmt["totals"].get(col, ""), col)
            cell.font = TOTAL_FONT
            cell.border = TOP_BORDER

    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"
    _autosize(ws, cols, [[r.get(c, "") for c in cols] for r in stmt["rows"]])


def _all_rows_sheet(wb: Workbook, statements: list[dict]) -> None:
    """Every row from every statement, mapped onto one cross-carrier schema."""
    ws = wb.create_sheet("All Rows", 0)
    headers = ["Source", "Carrier", *CANONICAL_FIELDS]
    ws.append(headers)
    _style_header(ws, len(headers))

    plain: list[list[str]] = []
    for stmt in statements:
        mapping = canonical_map(stmt.get("template", ""), stmt.get("columns", []))
        if not mapping:
            continue
        for row in stmt["rows"]:
            values = [row.get(mapping.get(f, ""), "") for f in CANONICAL_FIELDS]
            plain.append([stmt["filename"], stmt.get("template", ""), *values])
            ws.append([""] * len(headers))
            ws.cell(row=ws.max_row, column=1).value = stmt["filename"]
            ws.cell(row=ws.max_row, column=2).value = stmt.get("template", "")
            for i, (field_name, value) in enumerate(zip(CANONICAL_FIELDS, values), start=3):
                _write(ws.cell(row=ws.max_row, column=i), value, field_name)

    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    _autosize(ws, headers, plain)


def _checks_sheet(wb: Workbook, statements: list[dict]) -> None:
    ws = wb.create_sheet("Validation")
    headers = ["Source", "Carrier", "Check", "Rows total", "Stated total", "Status"]
    ws.append(headers)
    _style_header(ws, len(headers))

    plain: list[list[str]] = []
    # The failsafe first: it is the check that decides whether the numbers below
    # can be trusted at all.
    for stmt in statements:
        p = stmt.get("payout") or {}
        if not p:
            continue
        refs = "; ".join(f"{r['source']}: {r['amount']:,.2f}" for r in p.get("references", []))
        status = {
            "match": "FAILSAFE OK", "mismatch": "FAILSAFE FAILED",
            "no_reference": "FAILSAFE N/A", "no_amounts": "FAILSAFE N/A",
        }.get(p.get("status"), "FAILSAFE N/A")
        ws.append([
            stmt["filename"], stmt.get("template", ""),
            f"Exported {p.get('commission_column') or 'commission'} vs declared total",
            p.get("exported_total"), p.get("statement_total") or p.get("filename_total"),
            status if not refs else f"{status} ({refs})",
        ])
        for col in (4, 5):
            ws.cell(row=ws.max_row, column=col).number_format = MONEY_FMT
        ws.cell(row=ws.max_row, column=6).font = (
            OK_FONT if p.get("status") == "match" else BAD_FONT
        )
        ws.cell(row=ws.max_row, column=6).alignment = Alignment(wrap_text=True)
        plain.append([stmt["filename"], stmt.get("template", ""), "failsafe", "", "", status])

    for stmt in statements:
        for check in stmt.get("checks", []):
            status = "OK" if check["ok"] else "MISMATCH"
            ws.append(
                [
                    stmt["filename"],
                    stmt.get("template", ""),
                    check["label"],
                    check["rows_total"],
                    check["stated_total"],
                    status,
                ]
            )
            for col in (4, 5):
                ws.cell(row=ws.max_row, column=col).number_format = MONEY_FMT
            ws.cell(row=ws.max_row, column=6).font = OK_FONT if check["ok"] else BAD_FONT
            plain.append([stmt["filename"], stmt.get("template", ""), check["label"], "", "", status])

        for warning in stmt.get("warnings", []):
            ws.append([stmt["filename"], stmt.get("template", ""), "Warning", "", "", warning])
            ws.cell(row=ws.max_row, column=6).font = BAD_FONT
            ws.cell(row=ws.max_row, column=6).alignment = Alignment(wrap_text=True)
            plain.append([stmt["filename"], stmt.get("template", ""), "Warning", "", "", ""])

    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    _autosize(ws, headers, plain)
    ws.column_dimensions["F"].width = 44


def build_workbook(statements: list[dict]) -> BytesIO:
    """Return an in-memory .xlsx for the given parsed statements."""
    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet

    used: set[str] = set()
    for stmt in statements:
        _statement_sheet(wb, stmt, used)

    _checks_sheet(wb, statements)
    _all_rows_sheet(wb, statements)  # inserted at index 0, so it opens first

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
